from copy import copy
import openpyxl


class ExcelAccess:
    _book = None

    def __init__(self, file_path, data_only=False) -> None:
        self._book = openpyxl.load_workbook(file_path, data_only=data_only)

    def get_row(self, sheet, min_row=0):
        return self._book[sheet].iter_rows(min_row=min_row)

    def get_cell(self, sheet, row, col):
        return self._book[sheet].cell(row=row, column=col)

    def insert_cols(self, sheet, col_num, style_copy=False):
        self._book[sheet].insert_cols(col_num, 1)

        if style_copy is False:
            return

        rows = self._book[sheet].iter_rows()
        for row_num, row in enumerate(rows):
            to_cell = self._book[sheet].cell(row=row_num + 1, column=col_num)
            if row[col_num].has_style:
                to_cell._style = copy(row[col_num]._style)

    def edit_cell(self, sheet, row_num, col_num, value):
        cell = self._book[sheet].cell(row=row_num, column=col_num)
        cell.value = value

    def save(self, file_name):
        self._book.save(file_name)

    def close(self):
        self._book.close()
